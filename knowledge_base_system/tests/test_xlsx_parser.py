"""测试 XLSX 解析器 — 包括基础设施迁移、单次加载、链接细分、嵌入图片提取。"""

import io
import zipfile
from unittest.mock import patch

import pytest
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage

from app.core.models import AssetType, Document, ElementType
from parsers.xlsx_parser import XlsxParser


def _xlsx_bytes(workbook: Workbook) -> bytes:
    """将 openpyxl Workbook 序列化为字节。"""
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _parse(wb: Workbook, source_uri: str = "memory://xlsx") -> tuple:
    """辅助函数：创建 Document 并用 XlsxParser 解析。"""
    raw = _xlsx_bytes(wb)
    doc = Document(
        title="XLSX",
        source_type="xlsx",
        source_uri=source_uri,
        metadata={"raw_content": raw},
    )
    result = XlsxParser().parse(doc, raw)
    return result, doc


def _inject_formula_cache(raw: bytes, value: str) -> bytes:
    """向测试工作簿写入公式缓存值，用于覆盖 data_only 读取路径。"""
    input_buffer = io.BytesIO(raw)
    output_buffer = io.BytesIO()
    with zipfile.ZipFile(input_buffer, "r") as zin:
        with zipfile.ZipFile(output_buffer, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    text = data.decode("utf-8")
                    text = text.replace(
                        '<c r="B2"><f>SUM(A2:A3)</f><v></v></c>',
                        f'<c r="B2"><f>SUM(A2:A3)</f><v>{value}</v></c>',
                    )
                    data = text.encode("utf-8")
                zout.writestr(item, data)
    return output_buffer.getvalue()


class TestXlsxParser:
    """现有测试 — 验证重构后行为兼容。"""

    def setup_method(self):
        self.parser = XlsxParser()

    def test_supported_types(self):
        assert self.parser.supports("xlsx")
        assert self.parser.supports("XLSX")
        assert not self.parser.supports("xls")

    def test_parse_visible_sheets_and_skip_hidden(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "产品"
        ws1.append(["名称", "状态"])
        ws1.append(["知识库", "启用"])
        ws2 = wb.create_sheet("流程")
        ws2["A1"] = "步骤"
        ws2["A2"] = "上传"
        hidden = wb.create_sheet("隐藏")
        hidden.sheet_state = "hidden"
        hidden["A1"] = "不应解析"

        result, doc = _parse(wb)

        titles = [el for el in result.elements if el.element_type == ElementType.title]
        assert [title.text for title in titles] == ["产品", "流程"]
        assert all(title.source_location.section_path == [title.text] for title in titles)
        assert "隐藏" not in [el.text for el in result.elements]
        assert result.doc.source_hash.startswith("sha256:")
        assert all(el.doc_id == result.doc.doc_id for el in result.elements)

    def test_parse_simple_table_region(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "状态表"
        ws.append(["状态", "说明"])
        ws.append(["处理中", "系统正在解析文档"])
        ws.append(["成功", "文档已进入知识库"])

        result, _ = _parse(wb)
        tables = [el for el in result.elements if el.element_type == ElementType.table]
        # 2 个数据行 → 1 个元素，rows 包含全部数据行
        assert len(tables) == 1

        t = tables[0].structured_data["table"]
        assert t["headers"] == [{"text": "状态"}, {"text": "说明"}]
        assert len(t["rows"]) == 2
        assert t["rows"][0]["cells"][0]["text"] == "处理中"
        assert t["rows"][0]["cells"][0]["metadata"]["cell"] == "A2"
        assert t["rows"][1]["cells"][0]["text"] == "成功"
        assert t["rows"][1]["cells"][0]["metadata"]["cell"] == "A3"

    def test_split_multiple_table_regions(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "多表"
        ws["A1"] = "状态"
        ws["B1"] = "说明"
        ws["A2"] = "成功"
        ws["B2"] = "完成"
        ws["D1"] = "角色"
        ws["E1"] = "权限"
        ws["D2"] = "管理员"
        ws["E2"] = "全部"

        result, _ = _parse(wb)
        tables = [el for el in result.elements if el.element_type == ElementType.table]
        ranges = [table.structured_data["table"]["metadata"]["range"] for table in tables]

        assert ranges == ["A1:B2", "D1:E2"]

    def test_single_cell_region_becomes_paragraph(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "备注"
        ws["C3"] = "只是一条备注"

        result, _ = _parse(wb)
        paragraphs = [el for el in result.elements if el.element_type == ElementType.paragraph]

        assert len(paragraphs) == 1
        assert paragraphs[0].text == "只是一条备注"
        assert paragraphs[0].metadata["cell"] == "C3"
        assert paragraphs[0].metadata["sheet_name"] == "备注"

    def test_expand_merged_cells(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "合并"
        ws.merge_cells("A1:C1")
        ws["A1"] = "部门"
        ws["A2"] = "研发"
        ws["B2"] = "测试"
        ws["C2"] = "运维"

        result, _ = _parse(wb)
        table = next(el for el in result.elements if el.element_type == ElementType.table)
        table_data = table.structured_data["table"]

        assert table_data["headers"] == [{"text": "部门"}, {"text": "部门"}, {"text": "部门"}]

    def test_formula_cache_and_missing_cache_metadata(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "公式"
        ws["A1"] = "值"
        ws["B1"] = "合计"
        ws["A2"] = 1
        ws["A3"] = 2
        ws["B2"] = "=SUM(A2:A3)"

        # 缓存值存在
        cached_raw = _inject_formula_cache(_xlsx_bytes(wb), "3")
        cached_doc = Document(
            title="XLSX",
            source_type="xlsx",
            source_uri="memory://xlsx",
            metadata={"raw_content": cached_raw},
        )
        cached_result = self.parser.parse(cached_doc, cached_raw)
        cached_table = next(
            el for el in cached_result.elements if el.element_type == ElementType.table
        )
        cached_cell = cached_table.structured_data["table"]["rows"][0]["cells"][1]
        assert cached_cell["text"] == "3"
        assert cached_cell["metadata"]["formula"] == "=SUM(A2:A3)"
        assert cached_cell["metadata"]["formula_value_missing"] is False

        # 缓存值缺失
        missing_raw = _xlsx_bytes(wb)
        missing_doc = Document(
            title="XLSX",
            source_type="xlsx",
            source_uri="memory://xlsx",
            metadata={"raw_content": missing_raw},
        )
        missing_result = self.parser.parse(missing_doc, missing_raw)
        missing_table = next(
            el for el in missing_result.elements if el.element_type == ElementType.table
        )
        missing_cell = missing_table.structured_data["table"]["rows"][0]["cells"][1]
        assert missing_cell["text"] == "=SUM(A2:A3)"
        assert missing_cell["metadata"]["formula"] == "=SUM(A2:A3)"
        assert missing_cell["metadata"]["formula_value_missing"] is True

    def test_video_and_attachment_links_create_assets(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "链接"
        ws["A1"] = "类型"
        ws["B1"] = "地址"
        ws["A2"] = "视频"
        ws["B2"] = "https://example.com/demo.mp4"
        ws["A3"] = "附件"
        ws["B3"] = "说明书"
        ws["B3"].hyperlink = "https://example.com/manual.pdf"

        result, _ = _parse(wb)
        asset_types = {asset.original_uri: asset.asset_type for asset in result.assets}

        assert asset_types["https://example.com/demo.mp4"] == AssetType.video_link
        assert asset_types["https://example.com/manual.pdf"] == AssetType.document_link
        # 2 行合并为 1 个元素，2 个 Asset 都在同一元素的 asset_data 中
        tables = [el for el in result.elements if el.element_type == ElementType.table]
        assert len(tables) == 1
        assert len(tables[0].structured_data["table"]["rows"]) == 2
        # 所有 Asset 的 element_id 指向该元素
        for asset in result.assets:
            assert asset.element_id == tables[0].element_id

    def test_read_from_file_uri(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "备注"
        ws["A2"] = "来自文件"
        path = tmp_path / "sample.xlsx"
        path.write_bytes(_xlsx_bytes(wb))

        raw = path.read_bytes()
        doc = Document(
            title="File",
            source_type="xlsx",
            source_uri=f"file://{path}",
            metadata={"raw_content": raw},
        )
        result = self.parser.parse(doc, raw)

        assert result.doc.source_hash.startswith("sha256:")
        assert any(el.text == "来自文件" for el in result.elements)

    def test_invalid_workbook_raises_clear_error(self):
        raw = b"not a workbook"
        doc = Document(
            title="Invalid",
            source_type="xlsx",
            source_uri="memory://xlsx",
            metadata={"raw_content": raw},
        )
        with pytest.raises(ValueError, match="XLSX 解析失败"):
            self.parser.parse(doc, raw)


class TestLinkTypeClassification:
    """链接 URL 类型细分测试 — 任务 6.2 / 6.3 / 6.4。"""

    def setup_method(self):
        self.parser = XlsxParser()

    def test_image_link_classified_as_image(self):
        """图片链接被识别为 AssetType.image。"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "https://example.com/photo.png"
        ws["A2"] = "https://example.com/logo.jpg"
        ws["A3"] = "https://example.com/icon.gif"

        result, _ = _parse(wb)
        asset_types = {a.original_uri: a.asset_type for a in result.assets}

        assert asset_types["https://example.com/photo.png"] == AssetType.image_link
        assert asset_types["https://example.com/logo.jpg"] == AssetType.image_link
        assert asset_types["https://example.com/icon.gif"] == AssetType.image_link

    def test_video_link_classified_as_video(self):
        """视频链接被识别为 AssetType.video_link。"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "https://example.com/demo.mp4"
        ws["A2"] = "https://example.com/clip.webm"

        result, _ = _parse(wb)
        asset_types = {a.original_uri: a.asset_type for a in result.assets}

        assert asset_types["https://example.com/demo.mp4"] == AssetType.video_link
        assert asset_types["https://example.com/clip.webm"] == AssetType.video_link

    def test_document_link_classified_as_attachment(self):
        """文档链接被识别为 AssetType.document_link。"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "https://example.com/doc.pdf"
        ws["A2"] = "https://example.com/report.docx"

        result, _ = _parse(wb)
        asset_types = {a.original_uri: a.asset_type for a in result.assets}

        assert asset_types["https://example.com/doc.pdf"] == AssetType.document_link
        assert asset_types["https://example.com/report.docx"] == AssetType.document_link

    def test_image_link_via_hyperlink(self):
        """通过真实超链接设置的图片链接被识别为 AssetType.image。"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "产品图"
        ws["A1"].hyperlink = "https://example.com/product.png"

        result, _ = _parse(wb)
        assert result.assets[0].asset_type == AssetType.image_link
        assert result.assets[0].original_uri == "https://example.com/product.png"


class TestHyperlinkTextPreservation:
    """超链接文字保留测试 — 任务 6.5。"""

    def setup_method(self):
        self.parser = XlsxParser()

    def test_hyperlink_text_preserved(self):
        """超链接单元格的显示文字保留，不被 URL 覆盖。"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "说明书"
        ws["A1"].hyperlink = "https://example.com/manual.pdf"

        result, _ = _parse(wb)
        paragraph = next(
            el for el in result.elements if el.element_type == ElementType.paragraph
        )
        assert paragraph.text == "说明书"
        # URL 作为 Asset 提取
        assert len(result.assets) == 1
        assert result.assets[0].original_uri == "https://example.com/manual.pdf"

    def test_hyperlink_empty_text_fallback(self):
        """超链接单元格无显示文字时用占位符填充，Asset 记录原始 URL。"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = None
        ws["A1"].hyperlink = "https://example.com/manual.pdf"

        result, _ = _parse(wb)
        paragraph = next(
            el for el in result.elements if el.element_type == ElementType.paragraph
        )
        # 文本被占位符替换
        assert paragraph.text == "{{doc:1}}"
        # Asset 正确创建并通过 asset_data 关联
        assert len(result.assets) == 1
        assert result.assets[0].original_uri == "https://example.com/manual.pdf"
        assert result.assets[0].asset_type == AssetType.document_link
        assert paragraph.asset_data[0].asset_id == result.assets[0].asset_id


class TestMergedCellsWithHyperlink:
    """合并单元格 + 超链接组合场景 — 任务 6.6。"""

    def setup_method(self):
        self.parser = XlsxParser()

    def test_merged_cell_hyperlink_inherited(self):
        """合并区域中所有单元格继承源单元格的超链接。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "合并链接"
        ws.merge_cells("A1:B1")
        ws["A1"] = "产品手册"
        ws["A1"].hyperlink = "https://example.com/manual.pdf"
        # 需要 2 行 2 列才能被识别为表格区域
        ws["A2"] = "第二行"
        ws["B2"] = "数据"

        result, _ = _parse(wb)
        table = next(el for el in result.elements if el.element_type == ElementType.table)
        table_data = table.structured_data["table"]

        # 超链接被提取
        assert any(
            a.original_uri == "https://example.com/manual.pdf" for a in result.assets
        )

    def test_merged_cell_with_image_link(self):
        """合并单元格区域 + 图片链接。"""
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A1:B1")
        ws["A1"] = "Logo"
        ws["A1"].hyperlink = "https://example.com/logo.png"

        result, _ = _parse(wb)
        assert result.assets[0].asset_type == AssetType.image_link
        assert result.assets[0].original_uri == "https://example.com/logo.png"


class TestEmbeddedImageExtraction:
    """嵌入图片提取测试 — 任务 6.1。"""

    def setup_method(self):
        self.parser = XlsxParser()

    def test_embedded_image_extracted(self):
        """嵌入图片被提取为 AssetType.image 的 Asset。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "图片表"

        # 创建 1x1 像素的 PNG 图片
        img = XlImage(io.BytesIO(
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01"
            b"\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
            b"\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00\x00\x01\x01\x00\x05\x18\xd8N"
            b"\x00\x00\x00\x00IEND\xaeB`\x82"
        ))
        img.anchor = "B2"
        ws.add_image(img)

        result, _ = _parse(wb)
        image_assets = [a for a in result.assets if a.asset_type == AssetType.image]
        assert len(image_assets) == 1
        asset = image_assets[0]
        # 嵌入图片无外部来源，original_uri 为空；通过 _data 持有字节数据
        assert asset.original_uri == ""
        assert hasattr(asset, "_data")
        assert asset.metadata["source"] == "xlsx_image"
        assert asset.metadata["sheet_name"] == "图片表"
        assert hasattr(asset, "_data")

    def test_no_images_does_not_break(self):
        """无嵌入图片时不影响解析。"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "普通文本"

        result, _ = _parse(wb)
        image_assets = [a for a in result.assets if a.asset_type == AssetType.image]
        assert len(image_assets) == 0

    def test_image_asset_linked_to_cell_element(self):
        """嵌入图片 Asset 通过 asset_ids 关联到对应单元格的 ParsedElement。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "图片表"
        ws["A1"] = "标题"
        ws["B1"] = "图片"
        ws["A2"] = "内容"

        img = XlImage(io.BytesIO(
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01"
            b"\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
            b"\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00\x00\x01\x01\x00\x05\x18\xd8N"
            b"\x00\x00\x00\x00IEND\xaeB`\x82"
        ))
        img.anchor = "B2"  # 锚定到第二行第二列
        ws.add_image(img)

        result, _ = _parse(wb)
        table = next(el for el in result.elements if el.element_type == ElementType.table)

        # 图片 asset 的 asset_id 应该出现在表格的 asset_data 中
        image_asset = next(a for a in result.assets if a.asset_type == AssetType.image)
        assert any(ad.asset_id == image_asset.asset_id for ad in table.asset_data)
        # element_id 回链到表格元素
        assert image_asset.element_id == table.element_id


class TestGenericWebLinks:
    """普通网页链接测试 — 统一创建 web_link Asset，与 DOCX 解析器行为一致。"""

    def setup_method(self):
        self.parser = XlsxParser()

    def test_generic_web_link_creates_web_link_asset(self):
        """普通网页链接创建 AssetType.web_link 的 Asset。"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "参考链接"
        ws["A1"].hyperlink = "https://example.com/page"

        result, _ = _parse(wb)
        # 创建 1 个 web_link Asset
        assert len(result.assets) == 1
        asset = result.assets[0]
        assert asset.asset_type == AssetType.web_link
        assert asset.original_uri == "https://example.com/page"
        # 通过 asset_data 关联到对应元素
        paragraph = next(
            el for el in result.elements if el.element_type == ElementType.paragraph
        )
        assert len(paragraph.asset_data) == 1
        assert paragraph.asset_data[0].asset_id == asset.asset_id
        # element_id 回链
        assert asset.element_id == paragraph.element_id

    def test_mixed_resource_and_generic_links(self):
        """混合链接：全部创建 Asset，类型各不相同。"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "https://example.com/doc.pdf"
        ws["A2"] = "https://example.com/about"
        ws["A3"] = "https://youtube.com/watch?v=abc"

        result, _ = _parse(wb)
        # 3 个 URL 全部创建 Asset
        assert len(result.assets) == 3
        asset_types = {a.original_uri: a.asset_type for a in result.assets}
        assert asset_types["https://example.com/doc.pdf"] == AssetType.document_link
        assert asset_types["https://example.com/about"] == AssetType.web_link
        assert asset_types["https://youtube.com/watch?v=abc"] == AssetType.video_link
        # 通过 asset_data 关联到元素
        paragraphs = [el for el in result.elements if el.element_type == ElementType.paragraph]
        all_asset_ids = {ad.asset_id for el in paragraphs for ad in el.asset_data}
        assert all_asset_ids == {a.asset_id for a in result.assets}


class TestMultiSheetWithImagesAndLinks:
    """多 Sheet + 图片 + 链接组合场景 — 任务 6.7。"""

    def setup_method(self):
        self.parser = XlsxParser()

    def test_multi_sheet_with_mixed_resources(self):
        """多个 sheet 各自有链接和图片，资源正确归属。"""
        wb = Workbook()

        # Sheet 1: 视频链接
        ws1 = wb.active
        ws1.title = "视频"
        ws1["A1"] = "培训视频"
        ws1["A1"].hyperlink = "https://example.com/training.mp4"

        # Sheet 2: 图片链接
        ws2 = wb.create_sheet("图片")
        ws2["A1"] = "Logo"
        ws2["A1"].hyperlink = "https://example.com/logo.png"

        # Sheet 3: 文档链接
        ws3 = wb.create_sheet("文档")
        ws3["A1"] = "手册"
        ws3["A1"].hyperlink = "https://example.com/manual.pdf"

        result, _ = _parse(wb)

        # 三个 sheet 标题
        titles = [el.text for el in result.elements if el.element_type == ElementType.title]
        assert titles == ["视频", "图片", "文档"]

        # 资源类型正确
        asset_types = {a.original_uri: a.asset_type for a in result.assets}
        assert asset_types["https://example.com/training.mp4"] == AssetType.video_link
        assert asset_types["https://example.com/logo.png"] == AssetType.image_link
        assert asset_types["https://example.com/manual.pdf"] == AssetType.document_link

    def test_multi_sheet_with_embedded_images(self):
        """多个 sheet 各自有嵌入图片。"""
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "数据"

        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "图表"

        img = XlImage(io.BytesIO(
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01"
            b"\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
            b"\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00\x00\x01\x01\x00\x05\x18\xd8N"
            b"\x00\x00\x00\x00IEND\xaeB`\x82"
        ))
        img.anchor = "B2"
        ws2.add_image(img)

        result, _ = _parse(wb)
        titles = [el.text for el in result.elements if el.element_type == ElementType.title]
        assert titles == ["Sheet1", "Sheet2"]

        image_assets = [a for a in result.assets if a.asset_type == AssetType.image]
        assert len(image_assets) == 1
        assert image_assets[0].metadata["sheet_name"] == "Sheet2"
        assert image_assets[0].metadata["sheet_index"] == 2
