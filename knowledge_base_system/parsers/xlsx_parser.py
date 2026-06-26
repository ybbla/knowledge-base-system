"""XLSX 工作簿解析器
使用 openpyxl .xlsx 文件解析为统一ParsedElement Asset支持按区域自动检测表格与散列数据，处理合并单元格、超链接、公式和嵌入图片"""

import io
import logging
import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import PurePosixPath
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.exceptions import InvalidFileException

from app.core.models import (
    Asset,
    AssetData,
    AssetStatus,
    AssetType,
    Document,
    ElementType,
    ParsedElement,
    SourceLocation,
    compute_hash,
)
from parsers.base import DocumentParser, ParseResult, _BaseParseState
from parsers.utils import (
    is_attachment_url,
    is_video_url,
)

logger = logging.getLogger(__name__)

# 图片链接扩展名（docx_parser _IMAGE_EXTENSIONS 保持一致）
_IMAGE_EXTENSIONS: set[str] = {
    ".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg", ".tiff", ".tif",
}


@dataclass
class _CellInfo:
    """单元格信息，包含位置、文本、资源和元数据。"""
    row: int
    col: int
    coordinate: str
    text: str = ""
    asset_data: list[AssetData] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass
class _Region:
    """连续数据区域，用于区域检测和表格或段落判别。"""
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    @property
    def row_count(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def col_count(self) -> int:
        return self.max_col - self.min_col + 1

    @property
    def range_ref(self) -> str:
        start = f"{get_column_letter(self.min_col)}{self.min_row}"
        end = f"{get_column_letter(self.max_col)}{self.max_row}"
        return start if start == end else f"{start}:{end}"


class XlsxParser(DocumentParser):
    """XLSX 工作簿解析为统一 ParsedElement + Asset。

    支持 source_type：xlsx    逐工作表处理，多行多列区域视为表格，否则视为段落
    支持提取嵌入图片、超链接资源、合并单元格和公式。
    单元格内的 URL 用占位符替换（{{web:1}}），与 DOCX 行为一致。
    """

    SUPPORTED_TYPES = {"xlsx"}

    # 占位符前缀映射，与 DOCX/md 解析器保持一致
    _PLACEHOLDER_PREFIX = {
        "image": "image",
        "video": "video",
        "image_link": "image",
        "video_link": "video",
        "document_link": "doc",
        "web_link": "web",
    }

    def supports(self, source_type: str) -> bool:
        return source_type.lower() in self.SUPPORTED_TYPES

    # ── 主解析入──────────────────────────────────────────────────

    def parse(self, doc: Document, content: bytes | str) -> ParseResult:
        """将 XLSX 工作簿解析为结构化元素和资源列表。"""
        if isinstance(content, str):
            content = content.encode("utf-8")
        if not content:
            raise ValueError("XLSX 解析失败：文档内容为空")

        try:
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
        except TypeError as exc:
            # 部分 xlsx 文件的 styles.xml 中存在非标准 fill 定义（如 extLst 自定义填充），
            # openpyxl 反序列化时会抛出 TypeError。修复：从 ZIP 中删除有问题的 fill 元素后重试。
            wb = self._load_without_fill_styles(content, exc)
        except (InvalidFileException, OSError, KeyError, ValueError, zipfile.BadZipFile) as exc:
            raise ValueError(f"XLSX 解析失败：{exc}") from exc

        state = _XlsxParseState(doc.doc_id, doc.version)
        assets: list[Asset] = []
        assets_by_uri: dict[str, Asset] = {}
        # 占位符计数器（按类型前缀独立计数）和 URL→占位符 映射
        placeholder_counter: dict[str, int] = {}
        placeholder_map: dict[str, str] = {}

        for sheet_index, ws in enumerate(wb.worksheets, start=1):
            if ws.sheet_state != "visible":
                continue

            state.add_sheet_title(ws.title, sheet_index)

            # 先提取嵌入图片（在收集单元格之前，以便图asset_id 能合并到单元格）
            image_cell_map = self._extract_sheet_images(
                ws, doc, ws.title, sheet_index, assets, placeholder_counter,
            )

            # zip 预提取所有公式（无论缓存值是否存在）
            formulas_map = self._extract_all_formulas_from_zip(content, sheet_index)

            cells = self._collect_cells(
                ws, doc, sheet_index, assets, assets_by_uri,
                formulas_map=formulas_map, image_cell_map=image_cell_map,
                placeholder_counter=placeholder_counter, placeholder_map=placeholder_map,
            )
            for region in self._find_regions(cells):
                if region.row_count >= 2 and region.col_count >= 2:
                    state.add_table(ws.title, sheet_index, region, cells, assets_by_uri, assets)
                else:
                    state.add_paragraphs(ws.title, sheet_index, region, cells, assets_by_uri, assets)

        doc.source_hash = compute_hash(content)

        # 安全兜底：不应有无主 Asset，若存在则挂到第一个元素
        orphans = [a for a in assets if not a.element_id]
        if orphans and state.elements:
            logger.warning("xlsx 发现 %d 个无主 Asset，挂到首元素", len(orphans))
            first = state.elements[0]
            for asset in orphans:
                asset.element_id = first.element_id

        return ParseResult(doc=doc, elements=state.elements, assets=assets)

    # ── 嵌入图片提取 ────────────────────────────────────────────────

    def _extract_sheet_images(
        self,
        ws,
        doc: Document,
        sheet_name: str,
        sheet_index: int,
        assets: list[Asset],
        placeholder_counter: dict[str, int],
    ) -> dict[tuple[int, int], list[AssetData]]:
        """提取工作表中嵌入的图片，创建 Asset 并返单元格→AssetData 映射。

        占位符使用 {{image:N}} 与 DOCX 解析器保持一致，同一工作表内 N 递增。
        """
        image_cell_map: dict[tuple[int, int], list[AssetData]] = {}

        for idx, img in enumerate(ws._images):
            try:
                data = img._data()
            except Exception:
                continue

            # 将零基锚点转换为一基行列坐标。
            row, col = 0, 0
            try:
                anchor = img.anchor
                if hasattr(anchor, "_from"):
                    row = anchor._from.row + 1
                    col = anchor._from.col + 1
            except Exception:
                pass

            # 锚点提取失败（非标准锚定方式），无法确定图片属于哪个单元格，跳过
            if not row or not col:
                logger.warning("xlsx 图片锚点提取失败，跳过: sheet=%s idx=%d", sheet_name, idx)
                continue

            content_type = img.format or "png"
            ext = content_type.lower()

            asset = Asset(
                doc_id=doc.doc_id,
                asset_type=AssetType.image,
                original_uri="",                # 嵌入类型无外部来源
                status=AssetStatus.ready,
                metadata={
                    "source": "xlsx_image",
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_index,
                    "cell": f"{get_column_letter(col)}{row}",
                    "row": row,
                    "col": col,
                },
            )
            object.__setattr__(asset, "_data", data)
            assets.append(asset)

            prefix = "image"
            cnt = placeholder_counter.get(prefix, 0) + 1
            placeholder_counter[prefix] = cnt
            placeholder = f"{{{{{prefix}:{cnt}}}}}"
            image_cell_map.setdefault((row, col), []).append(
                AssetData(placeholder=placeholder, asset_id=asset.asset_id)
            )

        return image_cell_map

    # ── 单元格收──────────────────────────────────────────────────

    def _collect_cells(
        self,
        ws,
        doc: Document,
        sheet_index: int,
        assets: list[Asset],
        assets_by_uri: dict[str, Asset],
        formulas_map: dict[str, str],
        image_cell_map: dict[tuple[int, int], list[AssetData]] | None = None,
        placeholder_counter: dict[str, int] | None = None,
        placeholder_map: dict[str, str] | None = None,
    ) -> dict[tuple[int, int], _CellInfo]:
        """遍历工作表中的所有单元格，收集文本、公式、超链接和资源。

        仅处理真正的超链接对象（可 Ctrl+单击跳转），不提取纯文本 URL。
        """
        merged_map = self._build_merged_map(ws)
        cells: dict[tuple[int, int], _CellInfo] = {}
        if image_cell_map is None:
            image_cell_map = {}
        if placeholder_counter is None:
            placeholder_counter = {}
        if placeholder_map is None:
            placeholder_map = {}

        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                source_row, source_col = merged_map.get((row, col), (row, col))
                cell = ws.cell(row, col)
                source_cell = ws.cell(source_row, source_col)
                coordinate = cell.coordinate

                value_text = self._stringify(source_cell.value)
                hyperlink = self._hyperlink_target(cell) or self._hyperlink_target(source_cell)
                text = value_text
                metadata: dict[str, Any] = {
                    "cell": coordinate,
                    "sheet_name": ws.title,
                    "sheet_index": sheet_index,
                }

                # 从预提取 map 查询公式（无论缓存值是否存在）
                formula = formulas_map.get(source_cell.coordinate)
                if formula:
                    metadata["formula"] = formula
                    metadata["formula_value_missing"] = not bool(value_text)
                    if not value_text:
                        text = formula

                if (row, col) in merged_map and (row, col) != (source_row, source_col):
                    metadata["merged_from"] = source_cell.coordinate
                if hyperlink:
                    metadata["hyperlink"] = hyperlink
                    if not text and hyperlink.startswith(("http://", "https://")):
                        text = hyperlink

                # 提取可下载资源引用，URL 用占位符替换
                link_asset_data, text = self._assets_from_cell(
                    doc, ws.title, coordinate, text, hyperlink,
                    assets, assets_by_uri,
                    placeholder_counter, placeholder_map,
                )

                # 合并嵌入图片AssetData（图片优先排在前面）
                cell_asset_data = image_cell_map.get((row, col), []) + link_asset_data

                if not text and not cell_asset_data and not formula and not hyperlink:
                    continue

                metadata["text"] = text
                cells[(row, col)] = _CellInfo(
                    row=row,
                    col=col,
                    coordinate=coordinate,
                    text=text,
                    asset_data=cell_asset_data,
                    metadata=metadata,
                )

        return cells

    # ── 公式预提────────────────────────────────────────────────

    @staticmethod
    def _extract_all_formulas_from_zip(
        raw: bytes, sheet_index: int,
    ) -> dict[str, str]:
        """zip 原始 XML 中预提取当前工作表所有公式
        一次性解xl/worksheets/sheet{sheet_index}.xml        提取所<c r="..."><f>...</f></c> 节点        返回 {cell_ref: formula_text} 映射
        无论缓存值是否存在都提取公式文本        确保 metadata 中始终记录公式信息
        Args:
            raw: 原始 xlsx 文件字节            sheet_index: 工作表序号（1-based）
        Returns:
            {cell_ref: formula_text} 映射（如 {"B2": "=SUM(A2:A3)"}）        """
        formulas: dict[str, str] = {}
        try:
            sheet_xml_path = f"xl/worksheets/sheet{sheet_index}.xml"
            with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                if sheet_xml_path not in zf.namelist():
                    return formulas
                sheet_xml = zf.read(sheet_xml_path).decode("utf-8")

            # 匹配所<c r="B2"><f>SUM(A2:A3)</f> 节点
            for match in re.finditer(
                r'<c\s+r="([A-Z]+[0-9]+)"[^>]*>\s*<f[^>]*>(.*?)</f>',
                sheet_xml,
                re.DOTALL,
            ):
                cell_ref = match.group(1)
                formula_text = match.group(2).strip()
                formulas[cell_ref] = (
                    f"={formula_text}" if not formula_text.startswith("=") else formula_text
                )
        except Exception:
            pass

        return formulas

    # ── 合并单元格映──────────────────────────────────────────────

    @staticmethod
    def _build_merged_map(ws) -> dict[tuple[int, int], tuple[int, int]]:
        """构建合并单元格到源单元格的映射
        合并区域中所有单元格映射到左上角源单元格 (min_row, min_col)        """
        merged_map: dict[tuple[int, int], tuple[int, int]] = {}
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_map[(row, col)] = (min_row, min_col)
        return merged_map

    # ── 超链──────────────────────────────────────────────────────

    @staticmethod
    def _hyperlink_target(cell) -> str | None:
        """获取单元格的超链接目URL
        优先返回 hyperlink.target（外URL），
        其次返回 hyperlink.location（内部跳转）        """
        if cell.hyperlink is None:
            return None
        return cell.hyperlink.target or cell.hyperlink.location

    # ── 值序列化 ────────────────────────────────────────────────────

    @staticmethod
    def _stringify(value: Any) -> str:
        """将单元格值安全转换为字符串
        None 空字符串，datetime/date ISO 格式，其str.strip()        """
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ")
        if isinstance(value, date):
            return value.isoformat()
        return str(value).strip()

    # ── 链接资源提取 ────────────────────────────────────────────────

    def _assets_from_cell(
        self,
        doc: Document,
        sheet_name: str,
        coordinate: str,
        text: str,
        hyperlink: str | None,
        assets: list[Asset],
        assets_by_uri: dict[str, Asset],
        placeholder_counter: dict[str, int],
        placeholder_map: dict[str, str],
    ) -> tuple[list[AssetData], str]:
        """从单元格超链接中提取 URL 并创建 Asset。

        仅处理真正的超链接对象（可 Ctrl+单击跳转），不提取纯文本 URL。
        """
        if not hyperlink or not hyperlink.startswith(("http://", "https://")):
            return [], text

        url = hyperlink
        asset_data_list: list[AssetData] = []

        asset = assets_by_uri.get(url)
        if asset is None:
            asset_type = self._classify_link_asset_type(url)
            prefix = self._PLACEHOLDER_PREFIX.get(asset_type.value, "res")
            cnt = placeholder_counter.get(prefix, 0) + 1
            placeholder_counter[prefix] = cnt
            placeholder = f"{{{{{prefix}:{cnt}}}}}"
            placeholder_map[url] = placeholder
            display_text = text if text and text != url else ""
            asset = Asset(
                doc_id=doc.doc_id,
                asset_type=asset_type,
                original_uri=url,
                display_text=display_text,
                storage_uri=None,
                status=AssetStatus.ready,
                extracted_text=None,
                metadata={
                    "source": "xlsx_cell",
                    "sheet_name": sheet_name,
                    "cell": coordinate,
                },
            )
            assets.append(asset)
            assets_by_uri[url] = asset
        else:
            placeholder = placeholder_map.get(url, "")

        asset_data_list.append(
            AssetData(placeholder=placeholder, asset_id=asset.asset_id)
        )

        return asset_data_list, text

    @staticmethod
    def _classify_link_asset_type(url: str) -> AssetType:
        """根据 URL 判断链接的资源类型。

        优先级：视频 > 图片 > 附件 > 普通网页（兜底 web_link）。
        与 DOCX 解析器的 classify_link_text() 行为一致：所有链接统一创建 Asset。
        """
        if is_video_url(url):
            return AssetType.video_link

        suffix = PurePosixPath(url.split("?", 1)[0]).suffix.lower()
        if suffix in _IMAGE_EXTENSIONS:
            return AssetType.image_link

        if is_attachment_url(url):
            return AssetType.document_link

        # 普通网页链接，兜底为 web_link Asset
        return AssetType.web_link

    # ── 区域检────────────────────────────────────────────────────

    @staticmethod
    def _find_regions(cells: dict[tuple[int, int], _CellInfo]) -> list[_Region]:
        """从单元格集合中检测连续的数据区域
        使用 occupied_rows 映射（dict[int, set[int]]）快速跳过空候选区域，
        避免笛卡尔积产生的无效遍历        """
        if not cells:
            return []

        # 预构 row → 该行所有有数据的列的集合
        occupied_rows: dict[int, set[int]] = {}
        for row, col in cells:
            occupied_rows.setdefault(row, set()).add(col)

        row_groups = _group_contiguous(sorted(occupied_rows.keys()))
        all_cols = sorted({col for _, col in cells})
        col_groups = _group_contiguous(all_cols)
        regions: list[_Region] = []

        for row_start, row_end in row_groups:
            for col_start, col_end in col_groups:
                # O(1) 快速跳过空区域
                if not any(
                    any(col_start <= c <= col_end for c in occupied_rows.get(r, set()))
                    for r in range(row_start, row_end + 1)
                ):
                    continue
                regions.append(_Region(row_start, row_end, col_start, col_end))

        return sorted(regions, key=lambda region: (region.min_row, region.min_col))

    # ── 样式兼容性修复 ──────────────────────────────────────────────

    @staticmethod
    def _load_without_fill_styles(content: bytes, original_error: TypeError):
        """非标准 fill 样式导致 openpyxl 解析失败时，修复 styles.xml 后重新加载。

        部分 xlsx 文件的 styles.xml 中包含 openpyxl 无法反序列化的 fill 定义
        （如自定义扩展填充），反序列化时 Fill() 不接受参数导致 TypeError。
        修复策略：将 <fills> 中所有 fill 替换为等量的默认 patternFill，
        保持 fill 索引不变，避免 cellXfs 中的 fillId 引用越界。
        """
        import io as _io
        import zipfile as _zipfile

        try:
            with _zipfile.ZipFile(_io.BytesIO(content), "r") as zf_in:
                names = zf_in.namelist()
                if "xl/styles.xml" not in names:
                    raise ValueError(
                        f"XLSX 样式兼容性错误（{original_error}），"
                        "但文件中未找到 styles.xml，无法修复"
                    ) from original_error

                styles_xml = zf_in.read("xl/styles.xml").decode("utf-8")

                # 统计原有 fill 数量，每个替换为最简 patternFill（保持 count 和索引不变）
                fill_tags = re.findall(
                    r"<fill[^/>]*(?:/>|>.*?</fill>)", styles_xml, re.DOTALL,
                )
                fill_count = len(fill_tags)
                default_fills = (
                    f'<fills count="{fill_count}">'
                    + "".join(
                        ["<fill><patternFill patternType=\"none\"/></fill>"]
                        * fill_count
                    )
                    + "</fills>"
                )

                fill_fixed = re.sub(
                    r"<fills[^>]*>.*?</fills>",
                    default_fills,
                    styles_xml,
                    flags=re.DOTALL,
                )

                if fill_fixed == styles_xml:
                    raise ValueError(
                        f"XLSX 样式兼容性错误（{original_error}），"
                        "修复 styles.xml 后内容未变化"
                    ) from original_error

                # 重建 xlsx ZIP
                buffer = _io.BytesIO()
                with _zipfile.ZipFile(
                    buffer, "w", _zipfile.ZIP_DEFLATED
                ) as zf_out:
                    for name in names:
                        if name == "xl/styles.xml":
                            zf_out.writestr(name, fill_fixed.encode("utf-8"))
                        else:
                            zf_out.writestr(name, zf_in.read(name))

                return load_workbook(
                    _io.BytesIO(buffer.getvalue()),
                    data_only=True,
                    read_only=False,
                )

        except (ValueError, _zipfile.BadZipFile):
            raise
        except Exception as exc:
            raise ValueError(
                f"XLSX 样式兼容性错误（{original_error}），修复也失败：{exc}"
            ) from original_error

    # ── 清理 ──────────────────────────────────────────────────────

    @staticmethod
    def _cleanup_raw_content(doc: Document) -> None:
        """清理 metadata 中的 raw_content，避免大文件字节滞留内存。"""
        cleaned = dict(doc.metadata)
        cleaned.pop("raw_content", None)
        doc.metadata = cleaned

# ── 内部解析状────────────────────────────────────────────────────


@dataclass
class _XlsxParseState(_BaseParseState):
    """维护 XLSX 解析过程中生成元素的顺序和标题路径
    继承 _BaseParseState doc_id、doc_version、elements、_seq    _section_path _next_seq() 方法    """

    def add_sheet_title(self, sheet_name: str, sheet_index: int) -> None:
        """为每个工作表添加标题元素。"""
        self._section_path = [sheet_name]
        self.elements.append(
            ParsedElement(
                doc_id=self.doc_id,
                doc_version=self.doc_version,
                sequence_order=self._next_seq(),
                element_type=ElementType.title,
                text=sheet_name,
                source_location=SourceLocation(section_path=list(self._section_path)),
                metadata={
                    "heading_level": 1,
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_index,
                },
            )
        )

    # 每个表格元素最多包含的数据行数，超出按此分片
    _MAX_ROWS_PER_ELEMENT = 30

    @staticmethod
    def _strip_cell_meta(meta: dict) -> dict:
        """去掉 cell metadata 中与 table 级别重复的字段。"""
        return {
            k: v for k, v in meta.items()
            if k not in ("sheet_name", "sheet_index", "text")
        }

    def add_table(
        self,
        sheet_name: str,
        sheet_index: int,
        region: _Region,
        cells: dict[tuple[int, int], _CellInfo],
        assets_by_uri: dict[str, Asset],
        assets: list[Asset] | None = None,
    ) -> None:
        """将表格区域按 _MAX_ROWS_PER_ELEMENT 分片为多个 ParsedElement。

        首行为表头，每个元素复用同一表头 + 最多 30 行数据。
        合并单元格通过 _collect_cells 的 merged_map 已展开到各行。
        """
        # ── 表头（区域首行），格式对齐 DOCX：[{"text": "列A"}, ...] ──
        headers = [
            {"text": self._cell(cells, region.min_row, col).text}
            for col in range(region.min_col, region.max_col + 1)
        ]
        header_asset_data: list[AssetData] = []
        for col in range(region.min_col, region.max_col + 1):
            cell = self._cell(cells, region.min_row, col)
            header_asset_data.extend(cell.asset_data)

        # ── 收集所有数据行 ──
        all_rows: list[dict] = []
        for row in range(region.min_row + 1, region.max_row + 1):
            row_cells = []
            for col in range(region.min_col, region.max_col + 1):
                cell = self._cell(cells, row, col)
                row_cells.append(
                    {"text": cell.text, "metadata": self._strip_cell_meta(cell.metadata)}
                )
            all_rows.append({"cells": row_cells})

        # ── 按 _MAX_ROWS_PER_ELEMENT 分片创建元素 ──
        total_rows = len(all_rows)
        for start in range(0, total_rows, self._MAX_ROWS_PER_ELEMENT):
            chunk = all_rows[start : start + self._MAX_ROWS_PER_ELEMENT]
            first_row = region.min_row + 1 + start
            last_row = first_row + len(chunk) - 1

            # range 包含表头行，表示这个元素覆盖的完整区域
            chunk_range = _Region(
                region.min_row, last_row, region.min_col, region.max_col,
            )

            # 该分片涉及的行范围内的资源，去重注入
            chunk_asset_data = list(header_asset_data)
            seen_ids: set[str] = {ad.asset_id for ad in header_asset_data}
            for r in range(first_row, last_row + 1):
                for col in range(region.min_col, region.max_col + 1):
                    cell = self._cell(cells, r, col)
                    for ad in cell.asset_data:
                        if ad.asset_id not in seen_ids:
                            seen_ids.add(ad.asset_id)
                            chunk_asset_data.append(ad)

            structured = {
                "table": {
                    "caption": "",
                    "headers": headers,
                    "rows": chunk,
                    "metadata": {
                        "sheet_name": sheet_name,
                        "sheet_index": sheet_index,
                        "range": chunk_range.range_ref,
                    },
                }
            }
            text = self._table_text(headers, chunk)
            element = ParsedElement(
                doc_id=self.doc_id,
                doc_version=self.doc_version,
                sequence_order=self._next_seq(),
                element_type=ElementType.table,
                text=text,
                structured_data=structured,
                asset_data=chunk_asset_data,
                source_location=SourceLocation(
                    section_path=list(self._section_path),
                    table_path=[
                        {
                            "sheet_name": sheet_name,
                            "sheet_index": sheet_index,
                            "range": chunk_range.range_ref,
                        }
                    ],
                ),
            )
            self._link_assets(element, assets_by_uri, assets)
            self.elements.append(element)

    def add_paragraphs(
        self,
        sheet_name: str,
        sheet_index: int,
        region: _Region,
        cells: dict[tuple[int, int], _CellInfo],
        assets_by_uri: dict[str, Asset],
        assets: list[Asset] | None = None,
    ) -> None:
        """为单行或单列区域创建段落元素。"""
        for row in range(region.min_row, region.max_row + 1):
            for col in range(region.min_col, region.max_col + 1):
                cell = cells.get((row, col))
                if cell is None:
                    continue
                element = ParsedElement(
                    doc_id=self.doc_id,
                    doc_version=self.doc_version,
                    sequence_order=self._next_seq(),
                    element_type=ElementType.paragraph,
                    text=cell.text,
                    asset_data=list(cell.asset_data),
                    source_location=SourceLocation(
                        section_path=list(self._section_path),
                        table_path=[
                            {
                                "sheet_name": sheet_name,
                                "sheet_index": sheet_index,
                                "cell": cell.coordinate,
                            }
                        ],
                    ),
                    metadata={
                        **cell.metadata,
                        "sheet_name": sheet_name,
                        "sheet_index": sheet_index,
                    },
                )
                self._link_assets(element, assets_by_uri, assets)
                self.elements.append(element)

    @staticmethod
    def _cell(
        cells: dict[tuple[int, int], _CellInfo], row: int, col: int,
    ) -> _CellInfo:
        """安全获取单元格信息，缺失时返回空的 _CellInfo。"""
        coordinate = f"{get_column_letter(col)}{row}"
        return cells.get((row, col)) or _CellInfo(
            row=row, col=col, coordinate=coordinate,
        )

    @staticmethod
    def _table_text(headers: list[dict], rows: list[dict[str, Any]]) -> str:
        """将表格数据序列化为纯文本表示。"""
        parts: list[str] = []
        if headers:
            parts.append(" | ".join(h["text"] for h in headers))
        for row in rows:
            parts.append(" | ".join(cell["text"] for cell in row["cells"]))
        return "\n".join(part for part in parts if part.strip())

    @staticmethod
    def _link_assets(
        element: ParsedElement,
        assets_by_uri: dict[str, Asset],
        assets: list[Asset] | None = None,
    ) -> None:
        """通过 asset_id 将元素关联资源回填到 Asset.element_id。"""
        asset_ids = {ad.asset_id for ad in element.asset_data}
        for asset in assets_by_uri.values():
            if asset.asset_id in asset_ids and not asset.element_id:
                asset.element_id = element.element_id
        if assets:
            for asset in assets:
                if asset.asset_id in asset_ids and not asset.element_id:
                    asset.element_id = element.element_id


def _group_contiguous(values: list[int]) -> list[tuple[int, int]]:
    """将整数列表按连续性分组
    例：[1, 2, 3, 5, 6] [(1, 3), (5, 6)]
    """
    if not values:
        return []

    groups: list[tuple[int, int]] = []
    start = previous = values[0]
    for value in values[1:]:
        if value == previous + 1:
            previous = value
            continue
        groups.append((start, previous))
        start = previous = value
    groups.append((start, previous))
    return groups
